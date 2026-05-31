import tkinter as tk
import openpyxl as op
from tkinter import ttk,messagebox

# ===== FUNCTIONS =====
# Display Excel Data
def display():
    workbook = op.load_workbook("Felices_Database.xlsx")
    sheet = workbook.active
    
    for row in table.get_children():
        table.delete(row)
        
    for row in sheet.iter_rows(min_row=2, values_only=True):
    	table.insert("", tk.END, values=row)
    
    workbook.save("Felices_Database.xlsx")

# Check if the inputs entered is correct
def input_validation():
	product = product_entry.get()
	price = price_entry.get()
	quantity = quantity_entry.get()
	
	if not product or not price or not quantity:
		messagebox.showerror("Error", "Product, Price, and Quantity are required")
		return False
		
	if not price.isdigit() or not quantity.isdigit():
		messagebox.showerror("Error", "Price and Product must be a number")
		return False
		
	return True

# Add new record
def append():
	if not input_validation():
		return
		
	product = product_entry.get()
	price = price_entry.get()
	quantity = quantity_entry.get()
	total = int(price) * int(quantity)
	
	workbook = op.load_workbook("Felices_Database.xlsx")
	sheet = workbook.active
	
	new_id = sheet.max_row
	
	sheet.append([new_id, product, price, quantity, total])
	workbook.save("Felices_Database.xlsx")
	
	messagebox.showinfo("Success", "Record added")
	display()

# Select a Data
def auto_populate(event):
	selected = table.focus()
	values = table.item(selected, "values")
	
	if values:
		product_entry.delete(0, tk.END)
		price_entry.delete(0, tk.END)
		quantity_entry.delete(0, tk.END)
		
		product_entry.insert(0, values[1])
		price_entry.insert(0, values[2])
		quantity_entry.insert(0, values[3])

# Change Data
def update():
	selected = table.focus()
	
	if not selected:
		messagebox.showerror("Error", "Select a record first")
		return
		
	if not input_validation():
		return
	
	values = table.item(selected, "values")
	record_id = values[0]
	
	product = product_entry.get()
	price = price_entry.get()
	quantity = quantity_entry.get()
	total = int(price) * int(quantity)

	workbook = op.load_workbook("Felices_Database.xlsx")
	sheet = workbook.active
	
	for row in sheet.iter_rows(min_row=2):
		if str(row[0].value) == str(record_id):
			row[1].value = product
			row[2].value = price
			row[3].value = quantity
			
	workbook.save("Felices_Database.xlsx") 
	
	messagebox.showinfo("Success", "Updated successfully")
	display()

# Delete a Data
def delete():
	selected = table.focus()
	
	if not selected:
		messagebox.showerror("Error", "Select a record first")
		return
	
	values = table.item(selected, "values")
	record_id = values[0]
	
	confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this record?")
	if not confirm:
		return
		
	workbook = op.load_workbook("Felices_Database.xlsx")
	sheet = workbook.active
	
	for i, row in enumerate(sheet.iter_rows(min_row=2), start = 2):
		if str(row[0].value) == str(record_id):
			sheet.delete_rows(i)
			break
			
	workbook.save("Felices_Database.xlsx")
	
	messagebox.showinfo("Success", "Record Deleted Successfully")
	display()

# Hide Inputs - To only view the database
def hide():
	frame.grid_forget()
	hide_button['text'] = "> Show Inputs"
	hide_button['command'] = show

# Show Inputs
def show():
	frame.grid(row = 1, column = 0)
	hide_button['text'] = "< Hide Inputs"
	hide_button['command'] = hide

# ===== GUI =====
main_bg = "light blue"

window = tk.Tk()
window.title("Product Management System")
window.resizable(False, False)
window.config(bg = main_bg)

# Program Name
label = tk.Label(window, 
							text = "Product Management System", 
							font = ("Arial", 12, "bold"), 
							bg = main_bg)
label.grid(row = 0, column = 0, columnspan = 4)

# Frame
frame = tk.Frame(window, 
								bg = main_bg)
frame.grid(row = 1, column = 0)

# Product Name
product_label = tk.Label(frame, 
											text = "Product Name", 
											font = ("Arial", 8, "bold"), 
											bg = main_bg)
product_label.grid(row = 1, column = 0)

product_entry = tk.Entry(frame)
product_entry.grid(row = 0, column = 0)

# Price
price_label = tk.Label(frame, 
									text = "Price", 
									font = ("Arial", 8, "bold"), 
									bg = main_bg)
price_label.grid(row = 1, column = 1, padx = 10)

price_entry = tk.Entry(frame)
price_entry.grid(row = 0, column = 1, padx = 10)

# Quantity
quantity_label = tk.Label(frame, 
											text = "Quantity", 
											font = ("Arial", 8, "bold"), 
											bg = main_bg)
quantity_label.grid(row = 1, column = 2)

quantity_entry = tk.Entry(frame)
quantity_entry.grid(row = 0, column = 2)

# Submit
sub_button = tk.Button(frame, 
										text = "Submit", 
										font = ("Arial", 8, "bold"), 
										bg = "light green", 
										activebackground = "green", 
										command = append)
sub_button.grid(row = 2, column = 0)

# Update
upd_button = tk.Button(frame, 
										text = "Update", 
										font = ("Arial", 8, "bold"), 
										bg = "cyan", 
										activebackground = "dark cyan", 
										command = update)
upd_button.grid(row = 2, column = 1)

# Delete
del_button = tk.Button(frame, 
										text = "Delete", 
										font = ("Arial", 8, "bold"), 
										bg = "red", 
										activebackground = "dark red", 
										command = delete)
del_button.grid(row = 2, column = 2)

# Treeview - Database
table = ttk.Treeview(window,
									columns=("ID","Product", "Price", "Quantity", "Total Amount"),
									show="headings")
for headings in ("ID","Product","Price", "Quantity","Total Amount"):
    table.heading(headings,text=headings)
table.grid(row = 2, column = 0,columnspan=4, pady = 5)

# Hide Inputs
hide_button = tk.Button(window, 
										text = "< Hide Inputs", 
										font = ("Arial", 8, "bold"), 
										bg = "white", 
										activebackground = "gray", 
										command = hide)
hide_button.grid(row = 3, column = 0, columnspan = 4)

table.bind("<<TreeviewSelect>>", auto_populate)
display()

window.mainloop()